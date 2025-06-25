import os
from math import ceil, radians, cos, sin, asin, sqrt
from datetime import datetime
now = datetime.now()
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
import sqlite3
from app.database import get_db_connection


# Define file paths
BASE_DIR = os.getcwd()
UPLOAD_FOLDER = os.path.join(BASE_DIR, "uploads")
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
OUTPUT_FILE = os.path.join(UPLOAD_FOLDER, "output_kavach_slots_final_layout_v2.xlsx") # Updated output file name

def calculate_distance(lat1, lon1, lat2, lon2):
    from math import radians, cos, sin, asin, sqrt
    
    # Convert to radians
    lat1, lon1, lat2, lon2 = map(radians, [lat1, lon1, lat2, lon2])
    
    # Haversine formula
    dlat = lat2 - lat1
    dlon = lon2 - lon1
    a = sin(dlat/2)**2 + cos(lat1) * cos(lat2) * sin(dlon/2)**2
    c = 2 * asin(sqrt(a))
    km = 6371 * c  # Earth's radius in kilometers
    
    return km

# Helper function to parse timeslot string into a list of occupied indices (0-indexed)
def parse_timeslot_range(timeslot_str: str) -> set[int]:
    """
    Parses a timeslot string like '2-14' into a set of 0-indexed slot integers.
    'P2' corresponds to index 0, 'P45' corresponds to index 43.
    Assumes database format is 'START-END' (e.g., '2-45')
    """
    if not timeslot_str:
        return set()
    try:
        timeslot_str = timeslot_str.strip() # Remove any leading/trailing whitespace

        if '-' not in timeslot_str:
            # Handle single slot numbers if they might appear like "2" instead of "2-2"
            # Assuming if it's a single number, it still represents P_N
            num = int(timeslot_str)
            if 2 <= num <= 45: # P-numbers are from 2 to 45
                return set([num - 2])
            else:
                print(f"DEBUG: Warning: Single timeslot number '{timeslot_str}' is out of P2-P45 range. Returning empty set.")
                return set()
        
        start_num_str, end_num_str = timeslot_str.split('-')
        start_p = int(start_num_str)
        end_p = int(end_num_str)

        # Convert P-number to 0-indexed slot index (e.g., 2 -> 0, 45 -> 43)
        # P_N is index N-2
        # So, for '2-14': start_p_idx = 2-2=0, end_p_idx = 14-2=12
        # range(0, 12+1) which is range(0,13) -> [0, 1, ..., 12]
        
        # Ensure numbers are within valid P-number range before converting
        start_idx = max(0, start_p - 2) # P2 -> 0, anything below 2 still maps to 0 if relevant
        end_idx = min(43, end_p - 2)   # P45 -> 43, anything above 45 still maps to 43
        
        if start_idx > end_idx: # Edge case: if start P-number > end P-number (invalid range)
            print(f"DEBUG: Warning: Invalid timeslot range '{timeslot_str}'. Start index ({start_idx}) > End index ({end_idx}). Returning empty set.")
            return set()

        return set(range(start_idx, end_idx + 1)) # +1 because range end is exclusive
    except (ValueError, IndexError):
        print(f"DEBUG: Warning: Could not parse timeslot string '{timeslot_str}'. Expected format 'START-END' (e.g., '2-45') or single number ('2'). Returning empty set.")
        return set()

def calculate_all_overlaps(stations_list):
    overlapping_pair_data = []
    has_major_conflict = False

    for i in range(len(stations_list)):
        s1 = stations_list[i]
        for j in range(i + 1, len(stations_list)):
            s2 = stations_list[j]

            freq1 = int(s1.get('frequency', s1.get('allocated_frequency', 0)))
            freq2 = int(s2.get('frequency', s2.get('allocated_frequency', 0)))

            if freq1 == 0 or freq2 == 0 or freq1 != freq2:
                continue

            lat1 = float(s1.get('lat', s1.get('latitude', 0.0)))
            lon1 = float(s1.get('lon', s1.get('longitude', 0.0)))
            lat2 = float(s2.get('lat', s2.get('latitude', 0.0)))
            lon2 = float(s2.get('lon', s2.get('longitude', 0.0)))

            rad1 = float(s1.get('radius', s1.get('safe_radius_km', 0.0)))
            rad2 = float(s2.get('radius', s2.get('safe_radius_km', 0.0)))

            if any(val is None for val in [lat1, lon1, lat2, lon2, rad1, rad2]):
                print(f"Skipping overlap check due to missing/invalid coordinates/radius: {s1.get('name')} and {s2.get('name')}")
                continue

            dist_between_centers = calculate_distance(lat1, lon1, lat2, lon2)
            sum_of_radii = rad1 + rad2

            if dist_between_centers < sum_of_radii:
                has_major_conflict = True
                overlapping_pair_data.append({
                    's1_id': s1['id'],
                    's2_id': s2['id'],
                    's1_name': s1.get('name') or s1.get('stationName'),
                    's2_name': s2.get('name') or s2.get('stationName'),
                    'frequency': freq1,
                    'distance': dist_between_centers,
                    'min_required': sum_of_radii,
                    'line_coords': [[lat1, lon1], [lat2, lon2]],
                    'popup_content': f"FREQ {freq1} OVERLAP: {s1.get('name') or s1.get('stationName')} & {s2.get('name') or s2.get('stationName')} - Dist: {dist_between_centers:.2f}km (Req: {sum_of_radii:.2f}km)"
                })
    return overlapping_pair_data, has_major_conflict

def allocate_slots( 
    stations: list[dict],
    max_slots: int = 44,
    max_frequencies: int = 7
) -> list[dict]:
    """
    Allocates stationary and onboard slots using a relational 6-priority system,
    with an added layer of frequency-switching based on conflicts with approved stations'
    geographical coverage and localized timeslot usage.
    """
    allocations_output: list[dict] = []
    
    # This `frequency_slot_maps` tracks slots occupied by *planning stations*
    # `0` means free, `station_name` means occupied by that planning station.
    frequency_slot_maps = {
        f_id: {
            'station_alloc': [0] * max_slots, # `0` for free, station_name for occupied
            'onboard_alloc': [0] * max_slots  # `0` for free, station_name for occupied
        }
        for f_id in range(1, max_frequencies + 1)
    }

    print("DEBUG: Starting allocate_slots function.")
    # --- Fetch Approved Stations (only once) ---
    conn = get_db_connection()
    approved_stations = conn.execute("SELECT * FROM stations WHERE status = 'approved'").fetchall()
    conn.close()

    approved_stations_list = []
    print(f"DEBUG: Found {len(approved_stations)} approved stations in DB.")
    for ap_station_row in approved_stations:
        ap_station_dict = dict(ap_station_row)
        # Ensure numeric types are correct from DB
        ap_station_dict['allocated_frequency'] = int(ap_station_dict['allocated_frequency']) if ap_station_dict['allocated_frequency'] is not None else -1
        ap_station_dict['safe_radius_km'] = float(ap_station_dict['safe_radius_km']) if ap_station_dict['safe_radius_km'] is not None else 12.0
        ap_station_dict['latitude'] = float(ap_station_dict['latitude']) if ap_station_dict['latitude'] is not None else 0.0
        ap_station_dict['longitude'] = float(ap_station_dict['longitude']) if ap_station_dict['longitude'] is not None else 0.0
        
        # Parse timeslot range for approved stations if needed for future rule, though not directly used for blocking slots globally now.
        ap_station_dict['parsed_timeslots'] = parse_timeslot_range(ap_station_dict.get('timeslot', ''))

        approved_stations_list.append(ap_station_dict)
        print(f"DEBUG: Loaded Approved: {ap_station_dict['name']} (Freq: {ap_station_dict['allocated_frequency']}, Radius: {ap_station_dict['safe_radius_km']}km, Timeslot DB: '{ap_station_dict.get('timeslot', 'N/A')}' -> Parsed Indices: {ap_station_dict['parsed_timeslots']})")

    
    for station_data in stations:
        # --- Station Data Initialization (Using keys from your original code snippet) ---
        station_name = station_data["name"]
        optimum_static_param = station_data["Static"]
        requested_onboard_slots = station_data["onboardSlots"]
        station_code = station_data["StationCode"]
        skavach_id = station_data["KavachID"]
        latitude = station_data["Latitude"]
        longitude = station_data["Longitude"]
        # Assuming safe_radius_km is also part of planning station input, or use default
        safe_radius_km = station_data.get("SafeRadius", 12.0) # Use .get() for optional, with a default

        print(f"\nDEBUG: --- Processing Planning Station: {station_name} ---")
        print(f"DEBUG: Input Data for {station_name} - Lat: {latitude}, Lon: {longitude}, SafeRadius: {safe_radius_km}km, StaticParam: {optimum_static_param}, RequestedOnboard: {requested_onboard_slots}")

        # Input Type Checking / Validation
        if not all(isinstance(val, (int, float)) for val in [latitude, longitude, safe_radius_km, optimum_static_param, requested_onboard_slots]):
            print(f"DEBUG: Error for {station_name}: Invalid type for numeric input data. Lat:{type(latitude)}, Lon:{type(longitude)}, etc.")
            allocations_output.append({
                "Station": station_name, "Frequency": "N/A",
                "Error": "Invalid numeric input data types for calculation."
            })
            continue

        try:
            latitude = float(latitude)
            longitude = float(longitude)
            safe_radius_km = float(safe_radius_km)
            optimum_static_param = float(optimum_static_param)
            requested_onboard_slots = int(requested_onboard_slots)
        except (ValueError, TypeError) as e:
            print(f"DEBUG: Error for {station_name}: Type conversion failed for station data: {e}. Skipping.")
            allocations_output.append({
                "Station": station_name, "Frequency": "N/A",
                "Error": f"Type conversion failed for station data: {e}"
            })
            continue # Skip to next station

        val_for_roundup = ((optimum_static_param * 120) + (requested_onboard_slots - optimum_static_param) * 40 + 100) / 66
        calculated_station_slots_needed = ceil(val_for_roundup) if val_for_roundup > 0 else 1 # Ensure at least 1 slot if needed
        print(f"DEBUG: Calculated Stationary Slots Needed for {station_name}: {calculated_station_slots_needed}")

        committed_plan_details_for_station = None

        # --- Find a Suitable Frequency ---
        for current_freq_id_attempt in range(1, max_frequencies + 1):
            print(f"DEBUG: Attempting Frequency: {current_freq_id_attempt} for {station_name}")
            
            station_alloc_map = list(frequency_slot_maps[current_freq_id_attempt]['station_alloc']) # Use a copy for current checks
            onboard_alloc_map = list(frequency_slot_maps[current_freq_id_attempt]['onboard_alloc']) # Use a copy for current checks

            # --- 1. Geographical Conflict Check with Approved Stations on this frequency ---
            has_geo_conflict_with_approved = False 
            conflicting_approved_stations = [] 

            for ap_station in approved_stations_list:
                ap_freq = ap_station.get('allocated_frequency')
                
                # Only check if the approved station is on the current frequency attempt
                if ap_freq == current_freq_id_attempt:
                    distance = calculate_distance(latitude, longitude, ap_station['latitude'], ap_station['longitude'])
                    min_distance_sum = (safe_radius_km + ap_station['safe_radius_km'])
                    
                    if distance < min_distance_sum:
                        has_geo_conflict_with_approved = True
                        conflicting_approved_stations.append(f"{ap_station['name']} (Dist: {distance:.2f}km, MinReq: {min_distance_sum:.2f}km, AppFreq: {ap_freq})")
                        print(f"DEBUG:   -- Geo Conflict Detected: {station_name} (Planning) vs {ap_station['name']} (Approved) on Freq {current_freq_id_attempt}. Distance: {distance:.2f}km, Min Required: {min_distance_sum:.2f}km.")
                        break # Found one, this frequency is unsuitable based on your rule.

            if has_geo_conflict_with_approved:
                print(f"DEBUG: Freq {current_freq_id_attempt} is unsuitable for {station_name} due to geographical conflict with approved stations: {', '.join(conflicting_approved_stations)}. Trying next frequency.")
                continue # Skip to the next frequency

            print(f"DEBUG: Freq {current_freq_id_attempt} is geographically clear for {station_name} with all approved stations.")

            # --- 2. Find Prospective Stationary Slots (considering *only* previously allocated planning stations) ---
            # This logic remains the same as your original, operating on station_alloc_map which tracks PLANNING stations.
            prospective_station_slots_indices: list[int] = []
            
            for i in range(max_slots):
                if len(prospective_station_slots_indices) < calculated_station_slots_needed:
                    if station_alloc_map[i] == 0: # Check if slot is free (0) for planning stations
                        prospective_station_slots_indices.append(i)
                else: break # Enough slots found
            
            if len(prospective_station_slots_indices) < calculated_station_slots_needed:
                print(f"DEBUG: Not enough free stationary slots ({len(prospective_station_slots_indices)}/{calculated_station_slots_needed}) in Freq {current_freq_id_attempt} for {station_name} (taken by other planning stations). Trying next frequency.")
                continue # Not enough slots, try next frequency
            
            print(f"DEBUG: Found {len(prospective_station_slots_indices)} prospective stationary slots for {station_name} on Freq {current_freq_id_attempt}: {sorted([s+2 for s in prospective_station_slots_indices])}")


            # --- Step 2 (cont.): Plan-Then-Reclassify Onboard Slots ---
            # This part of the logic remains largely the same as your original code,
            # using the `station_alloc_map` and `onboard_alloc_map` which
            # reflect commitments from *other planning stations*.
            
            # Non-Stationary Planning (Proto-P1 and P3)
            planned_proto_p1: list[int] = []
            planned_p3: list[int] = []
            set_prospective_station_slots = set(prospective_station_slots_indices)

            if requested_onboard_slots > 0:
                idx = 0
                while len(planned_proto_p1) < requested_onboard_slots and idx < max_slots:
                    if (onboard_alloc_map[idx] == 0 and idx not in set_prospective_station_slots):
                        is_adjacent_to_this_planning_stationary = (idx > 0 and idx - 1 in set_prospective_station_slots) or \
                                                            (idx < max_slots - 1 and idx + 1 in set_prospective_station_slots)
                        
                        if not is_adjacent_to_this_planning_stationary:
                            planned_proto_p1.append(idx)
                            idx += 2 
                        else:
                            idx += 1 
                    else:
                        idx += 1 
                
            onboard_to_place_p3 = requested_onboard_slots - len(planned_proto_p1)
            if onboard_to_place_p3 > 0:
                for i in range(max_slots):
                    if len(planned_p3) < onboard_to_place_p3:
                        if (onboard_alloc_map[i] == 0 and i not in set_prospective_station_slots and i not in planned_proto_p1):
                            planned_p3.append(i)
                    else: break
            
            # Re-classify Proto-P1 into final P1 and P2
            planned_p1: list[int] = []
            planned_p2: list[int] = []
            set_p3 = set(planned_p3)
            for slot in planned_proto_p1:
                is_adjacent_to_p3 = (slot - 1 in set_p3) or (slot + 1 in set_p3)
                if is_adjacent_to_p3:
                    planned_p2.append(slot)
                else:
                    planned_p1.append(slot)
            
            # On-Stationary Planning (Proto-P4 and P6)
            planned_proto_p4: list[int] = []
            planned_p6: list[int] = []
            # Filter candidates: must be one of the *current planning station's* prospective stationary slots,
            # and free for onboard within the temporary map.
            on_stationary_candidates = [idx for idx in prospective_station_slots_indices 
                                        if onboard_alloc_map[idx] == 0 
                                        and idx not in (planned_p1 + planned_p2 + planned_p3)] # Ensure it's not already designated as non-stationary onboard

            if on_stationary_candidates:
                bottom_most_candidate = on_stationary_candidates[0]
                main_p4_candidates = on_stationary_candidates[1:]
                
                # P4a: Alternating
                for i in range(0, len(main_p4_candidates), 2):
                    if len(planned_p1) + len(planned_p2) + len(planned_p3) + len(planned_proto_p4) >= requested_onboard_slots: break
                    planned_proto_p4.append(main_p4_candidates[i])
                
                # P4b: Take bottom-most
                if len(planned_p1) + len(planned_p2) + len(planned_p3) + len(planned_proto_p4) < requested_onboard_slots:
                    if bottom_most_candidate not in planned_proto_p4:
                        planned_proto_p4.append(bottom_most_candidate)
                
                # Plan P6 (Continuous fill of on-stationary gaps)
                p6_candidates = [c for c in on_stationary_candidates if c not in planned_proto_p4]
                for c in p6_candidates:
                    if len(planned_p1) + len(planned_p2) + len(planned_p3) + len(planned_proto_p4) + len(planned_p6) >= requested_onboard_slots: break
                    planned_p6.append(c)

            # Re-classify Proto-P4 into final P4 and P5
            planned_p4: list[int] = []
            planned_p5: list[int] = []
            set_p6 = set(planned_p6)
            for slot in planned_proto_p4:
                is_adjacent_to_p6 = (slot - 1 in set_p6) or (slot + 1 in set_p6)
                if is_adjacent_to_p6:
                    planned_p5.append(slot)
                else:
                    planned_p4.append(slot)

            slot_p45_idx = 43
            all_lists = {
                'p1': planned_p1, 'p2': planned_p2, 'p4': planned_p4, 
                'p5': planned_p5, 'p6': planned_p6
            }
            for key, p_list in all_lists.items():
                if slot_p45_idx in p_list:
                    p_list.remove(slot_p45_idx)
                    if slot_p45_idx not in planned_p3:
                        planned_p3.append(slot_p45_idx)
                    print(f"DEBUG: P45 (slot {slot_p45_idx+2}) moved from {key} to P3 for {station_name}.")
                    break
            
            # --- Finalize Plan for this Frequency Attempt ---
            total_onboard_planned = len(planned_p1) + len(planned_p2) + len(planned_p3) + len(planned_p4) + len(planned_p5) + len(planned_p6)
            all_onboard_met = (total_onboard_planned >= requested_onboard_slots)

            if all_onboard_met: 
                committed_plan_details_for_station = {
                    "station_name": station_name, "frequency": current_freq_id_attempt,
                    "calculated_station_slots": calculated_station_slots_needed,
                    "prospective_station_slots_indices": list(prospective_station_slots_indices),
                    "requested_onboard_slots": requested_onboard_slots,
                    "onboard_indices_p1": list(planned_p1), "onboard_indices_p2": list(planned_p2),
                    "onboard_indices_p3": list(planned_p3), "onboard_indices_p4": list(planned_p4),
                    "onboard_indices_p5": list(planned_p5), "onboard_indices_p6": list(planned_p6),
                    "Static": optimum_static_param, "StationCode": station_code, 
                    "KavachID": skavach_id, "Latitude": latitude, "Longitude": longitude,
                    "SafeRadius": safe_radius_km # Add SafeRadius to committed plan
                }
                print(f"DEBUG: Successfully planned {total_onboard_planned} onboard slots (requested {requested_onboard_slots}) for {station_name} on Freq {current_freq_id_attempt}. Committing plan.")
                break # Found a suitable frequency, break from frequency loop
            else:
                print(f"DEBUG: Failed to allocate all requested onboard slots ({total_onboard_planned}/{requested_onboard_slots}) for {station_name} on Freq {current_freq_id_attempt}. Trying next frequency.")
        
        # --- Step 4: Commit the Chosen Plan if found ---
        if committed_plan_details_for_station:
            chosen_freq = committed_plan_details_for_station["frequency"]
            s_name_commit = committed_plan_details_for_station["station_name"]

            stat_p_nums_allocated: list[str] = []
            # Mark stationary slots as occupied by this planning station in the *global* map
            for slot_idx in committed_plan_details_for_station["prospective_station_slots_indices"]:
                frequency_slot_maps[chosen_freq]['station_alloc'][slot_idx] = s_name_commit
                stat_p_nums_allocated.append(f"P{slot_idx+2}")

            onboard_p_nums_overall: list[str] = []
            onboard_p1_nums, onboard_p2_nums, onboard_p3_nums, onboard_p4_nums, onboard_p5_nums, onboard_p6_nums = [], [], [], [], [], []
            
            all_planned_indices_tuples = [
                (onboard_p1_nums, committed_plan_details_for_station["onboard_indices_p1"]),
                (onboard_p2_nums, committed_plan_details_for_station["onboard_indices_p2"]),
                (onboard_p3_nums, committed_plan_details_for_station["onboard_indices_p3"]),
                (onboard_p4_nums, committed_plan_details_for_station["onboard_indices_p4"]),
                (onboard_p5_nums, committed_plan_details_for_station["onboard_indices_p5"]),
                (onboard_p6_nums, committed_plan_details_for_station["onboard_indices_p6"])
            ]

            current_onboard_placed_count = 0
            for p_num_list, p_indices_list in all_planned_indices_tuples:
                for slot_idx in sorted(list(set(p_indices_list))): # Use set to remove duplicates before sorting
                    if current_onboard_placed_count < committed_plan_details_for_station["requested_onboard_slots"]:
                        # Check if slot is truly free (0) in the global map for onboard allocations
                        if frequency_slot_maps[chosen_freq]['onboard_alloc'][slot_idx] == 0: 
                            frequency_slot_maps[chosen_freq]['onboard_alloc'][slot_idx] = s_name_commit
                            p_num = f"P{slot_idx+2}"
                            p_num_list.append(p_num)
                            onboard_p_nums_overall.append(p_num)
                            current_onboard_placed_count += 1
                        else:
                            print(f"DEBUG: Warning: Onboard slot P{slot_idx+2} in Freq {chosen_freq} already occupied by {frequency_slot_maps[chosen_freq]['onboard_alloc'][slot_idx]} (another planning station) when placing {s_name_commit}'s onboard slots. Skipping this slot for {s_name_commit}.")
                    else: break
            
            # Determine the allocated timeslot for this planning station
            allocated_stationary_indices_sorted = sorted(committed_plan_details_for_station["prospective_station_slots_indices"])
            allocated_timeslot_str = "N/A" # Default
            if allocated_stationary_indices_sorted:
                min_idx_p = allocated_stationary_indices_sorted[0] + 2 # Convert back to P-number
                max_idx_p = allocated_stationary_indices_sorted[-1] + 2 # Convert back to P-number
                allocated_timeslot_str = f"P{min_idx_p}-P{max_idx_p}" # Store as Px-Py format

            print(f"DEBUG: {s_name_commit} successfully allocated on Frequency {chosen_freq}. Stationary Slots: {', '.join(sorted(stat_p_nums_allocated, key=lambda x: int(x[1:])))}. Onboard Slots: {', '.join(sorted(onboard_p_nums_overall, key=lambda x: int(x[1:])))}. Allocated Timeslot Range: {allocated_timeslot_str}")

            allocations_output.append({
                "Station": s_name_commit, "Frequency": chosen_freq,
                "Stationary Kavach ID": committed_plan_details_for_station["KavachID"],
                "Station Code": committed_plan_details_for_station["StationCode"],
                "Latitude": committed_plan_details_for_station["Latitude"],
                "Longitude": committed_plan_details_for_station["Longitude"],
                "SafeRadius": committed_plan_details_for_station["SafeRadius"], # Include SafeRadius in output
                "Static": optimum_static_param,
                "Stationary Kavach Slots Requested": committed_plan_details_for_station["calculated_station_slots"],
                "Stationary Kavach Slots Allocated": ", ".join(sorted(stat_p_nums_allocated, key=lambda x: int(x[1:]))),
                "Num Stationary Allocated": len(stat_p_nums_allocated),
                "Onboard Kavach Slots Requested": committed_plan_details_for_station["requested_onboard_slots"],
                "Onboard Kavach Slots Allocated": ", ".join(sorted(onboard_p_nums_overall, key=lambda x: int(x[1:]))),
                "Num Onboard Allocated": len(onboard_p_nums_overall),
                "Onboard Slots P1 Allocated": ", ".join(sorted(onboard_p1_nums, key=lambda x: int(x[1:]))),
                "Onboard Slots P2 Allocated": ", ".join(sorted(onboard_p2_nums, key=lambda x: int(x[1:]))),
                "Onboard Slots P3 Allocated": ", ".join(sorted(onboard_p3_nums, key=lambda x: int(x[1:]))),
                "Onboard Slots P4 Allocated": ", ".join(sorted(onboard_p4_nums, key=lambda x: int(x[1:]))),
                "Onboard Slots P5 Allocated": ", ".join(sorted(onboard_p5_nums, key=lambda x: int(x[1:]))),
                "Onboard Slots P6 Allocated": ", ".join(sorted(onboard_p6_nums, key=lambda x: int(x[1:]))),
                "Allocated Timeslot Range": allocated_timeslot_str
            })
        else:
            print(f"DEBUG: Failed to find any suitable frequency for {station_name} after checking all {max_frequencies} frequencies. Marking as N/A.")
            allocations_output.append({
                "Station": station_name, "Frequency": "N/A",
                "Stationary Kavach ID": skavach_id, "Station Code": station_code,
                "Latitude": latitude, "Longitude": longitude, "Static": optimum_static_param,
                "Stationary Kavach Slots Requested": calculated_station_slots_needed,
                "Stationary Kavach Slots Allocated": "", "Num Stationary Allocated": 0,
                "Onboard Kavach Slots Requested": requested_onboard_slots,
                "Onboard Kavach Slots Allocated": "", "Num Onboard Allocated": 0,
                "Onboard Slots P1 Allocated": "", "Onboard Slots P2 Allocated": "", 
                "Onboard Slots P3 Allocated": "", "Onboard Slots P4 Allocated": "",
                "Onboard Slots P5 Allocated": "", "Onboard Slots P6 Allocated": "",
                "Error": "No suitable frequency found due to geographical conflict with approved stations or insufficient available slots for planning."
            })
            
    print("DEBUG: allocate_slots function finished.")
    return allocations_output

def generate_excel(input_stations_data):
    alloc_results = allocate_slots(input_stations_data)
    print("DEBUG processing.py: \n Generating Excel file with new styling logic...")

    try:
        df = pd.DataFrame(alloc_results)
        expected_cols = [
            "Station", "Static", "Frequency",
            "Stationary Kavach Slots Requested", "Stationary Kavach Slots Allocated", "Num Stationary Allocated",
            "Onboard Kavach Slots Requested", "Onboard Kavach Slots Allocated", "Num Onboard Allocated",
            "Onboard Slots P1 Allocated", "Onboard Slots P2 Allocated", "Onboard Slots P3 Allocated", "Onboard Slots P4 Allocated", 
            "Onboard Slots P5 Allocated", "Onboard Slots P6 Allocated",
            "Debug_IsIdeal", "Debug_Congested", "Debug_ExcessiveP3", "Error"
        ]
        for col in expected_cols:
            if col not in df.columns:
                df[col] = ""
        
        apply_color_scheme(df) # Pass the full DataFrame

        if os.path.exists(OUTPUT_FILE):
            print(f"Excel file saved successfully: {OUTPUT_FILE}")
            return OUTPUT_FILE
        else:
            print(f"Warning: {OUTPUT_FILE} was not created.")
            # Fallback to save uncolored if coloring somehow fails to save
            df.to_excel(os.path.join(UPLOAD_FOLDER,"fallback_uncolored_results.xlsx"), index=False)
            return None

    except Exception as e:
        print(f"Error generating Excel file: {e}")
        import traceback
        traceback.print_exc()
        return None

def  apply_color_scheme(results_df: pd.DataFrame): # Using user's function name
    if results_df.empty:
        print("Error: Input DataFrame for coloring is empty.")
        return

    now = datetime.now() # For use in titles

    color_map = {
        1: "F0F005", 2: "8FCA1D", 3: "F39D1B", 4: "3197EA",
        5: "90918F", 6: "F53B3D", 7: "CC6CE7"
    }
    # Corrected font colors to ARGB format
    FONT_P1_STYLE = Font(color="FF007220")  # Green
    FONT_P2_STYLE = Font(bold=True, color="FF0000FF")  # Blue Bold
    FONT_P3_STYLE = Font(color="FFE4080A")  # Red
    FONT_P4_STYLE = Font(bold=True, color="FF000000") # Black Bold
    FONT_P5_STYLE = Font(bold=True, color="FF000000", underline='single')  # Black Bold Underlined
    FONT_P6_STYLE = Font(color="FFFFFF", bold=True)  # White Bold

    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                         top=Side(style='thin'), bottom=Side(style='thin'))
    
    left_align_v_center_wrap = Alignment(horizontal='left', vertical='center', wrap_text=True)
    center_align_v_center_wrap = Alignment(horizontal='center', vertical='center', wrap_text=True)
    center_align_v_center = Alignment(horizontal='center', vertical='center')
    center_align_v_center_no_wrap = Alignment(text_rotation=90,horizontal='center', vertical='center', wrap_text=False)
    left_align_v_center = Alignment(horizontal='left', vertical='center') # For blank data cells

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Slot Allocation Matrix"

    # --- Define Row Numbers for Content (from user's code) ---
    title_row1 = 2
    title_row2 = 3
    title_row3 = 4
    title_row4 = 5
    # Row 6 is blank
    label_stationary_kavach_id_row = 7
    header_excel_row = 8  # "Station Name" label in Col A, Rotated station names in Col B+
    label_station_code_row = 9
    label_lat_row = 10
    label_long_row = 11
    data_optimum_static_row = 12
    label_freq_pair_row = 13    # Row 13 for Frequency
    stationary_count_row = 14 
    label_tx_window_row = 15    # Row 15 for Tx Window (P{min}-P{max})
    onboard_count_row = 16    
    data_start_row = 17       # Px data (P2, P3...) starts in Col A

    all_slots = [f"P{i}" for i in range(2, 46)] 
    all_stations = results_df["Station"].unique()
    max_excel_col = len(all_stations) + 1 
    max_slot_idx_for_adj_check = len(all_slots) - 1 # Renamed from max_slot_index for clarity

    # --- Write Titles (Rows 2-5) ---
    titles_config = [
        (title_row1, "Kavach (TCAS) : Application-cum-Approval: mComm Frequency Channels & Timeslots"),
        (title_row2, "Stationary Kavach Unit-wise Frequency Channels - Timeslot Details"),
        (title_row3, f"Application Number: Kavach/mComm/Appl/NCR-to-CoE/003/{now.hour:02d}"),
        (title_row4, f"Station - Station (Excl): When in Category-C (Radio Packet Structure as well as Tag Data Foramt as per V4.0 with SRS 4.0d3 Annex-C Amdt-7 wef {now.strftime('%d-%m-%Y')})")
    ]
    for r_num, text in titles_config:
        cell = ws.cell(row=r_num, column=1)
        cell.value = text
        ws.merge_cells(start_row=r_num, start_column=1, end_row=r_num, end_column=max_excel_col)
        cell.alignment = center_align_v_center_wrap

     
    legend_col = 1
    legend_start_row = 61

    headers = ["Legend: Onboard Tx Slot Priorities", "Example"]
    legend_data = [
        ["Priority 1 Green", "007220", "P2, P4, P6"],
        ["Priority 2 Blue Bold", "0000FF", "P8, P10, P12"],
        ["Priority 3 Red", "E4080A", "P9, P11, P13"],
        ["Prority 4 Black Bold", "000000", "P20, P22, P24"],
        ["Priority 5 Black Underlined Bold", "000000", "P26, P28, P30"],
        ["Priority 6 White Bold", "FFFFFF", "P27, P29, P31"]
    ]
    # Define border style
    border_style = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    # Apply headers with border
    for i, header in enumerate(headers):
        cell = ws.cell(row=legend_start_row, column=legend_col + i, value=header)
        cell.font = Font(bold=True)
        cell.border = border_style
    for row_offset, (label, hex_color, example) in enumerate(legend_data, start=1):
        bold = label.endswith("Bold")  # Check if the label should be bold
        underlined = label.endswith("Underlined Bold")  # Check if the label should be underlined
        
        label_cell = ws.cell(row=legend_start_row + row_offset, column=legend_col, value=label)
        label_cell.font = Font(bold=bold, underline='single' if underlined else None, color=f"FF{hex_color}")  # Apply bold formatting conditionally
        
        example_cell = ws.cell(row=legend_start_row + row_offset, column=legend_col + 1, value=example)
        example_cell.font = Font(bold=bold, underline='single' if 'Priority 5' in label else None, color=f"FF{hex_color}")

        # Apply dark grey fill for Priority 4, 5, and 6
        if int(label.split()[1]) >= 4:
            example_cell.fill = PatternFill(start_color="A7A7A7", end_color='A7A7A7', fill_type='solid')
            label_cell.fill = PatternFill(start_color='A7A7A7', end_color='A7A7A7', fill_type='solid')
        
        example_cell.border = border_style
        label_cell.border = border_style

    # Adjust column width for better visibility
    ws.column_dimensions[ws.cell(row=legend_start_row, column=legend_col).column_letter].width = 36
    ws.column_dimensions[ws.cell(row=legend_start_row, column=legend_col + 1).column_letter].width = 10

    # --- Write Static Labels in Column A ---
    static_labels_col1 = {
        label_stationary_kavach_id_row: "Stationary Kavach ID",
        header_excel_row: "Station Name", 
        label_station_code_row: "Station code",
        label_lat_row: "Stationary Unit Tower Latitude",
        label_long_row: "Stationary Unit Tower Longitude",
        data_optimum_static_row: "Optimum no. of Simultaneous Exclusive Static Profile Transfer",
        label_freq_pair_row: "Proposed Frequency Pair",
        stationary_count_row: "Number of Stationary Kavach Tx slots",
        label_tx_window_row: "Stationary Kavach (TCAS) Tx Window Commence - End",
        onboard_count_row: "Peak nos. of Onboard Kavach Units in Stn Unit Jurisdiction"
    }
    for r_num, text in static_labels_col1.items():
        cell = ws.cell(row=r_num, column=1)
        cell.value = text
        # Apply specific alignment for col A labels
        if r_num == header_excel_row: # "Station Name" label
            cell.alignment = center_align_v_center
        else:
            cell.alignment = center_align_v_center

    # --- Populate Data for Specific Rows (Columns B onwards) ---
    for c_idx_df, station_name in enumerate(all_stations):
        excel_data_col = c_idx_df + 2 
        station_specific_data_rows = results_df[results_df["Station"] == station_name]
        if station_specific_data_rows.empty: continue
        station_data_row = station_specific_data_rows.iloc[0]

        # Row 7: Stationary Kavach ID
        cell_id_r7 = ws.cell(row=label_stationary_kavach_id_row, column=excel_data_col)
        cell_id_r7.value = station_data_row.get("Stationary Kavach ID", "")
        cell_id_r7.alignment = center_align_v_center_no_wrap


        # Row 8 (Rotated Station Names)
        header_cell_rotated = ws.cell(row=header_excel_row, column=excel_data_col)
        header_cell_rotated.value = station_name
        header_cell_rotated.alignment = center_align_v_center_no_wrap
        
        # Row 9: Station Code
        cell_code_r9 = ws.cell(row=label_station_code_row, column=excel_data_col)
        cell_code_r9.value = station_data_row.get("Station Code", "")
        cell_code_r9.alignment = center_align_v_center_no_wrap

        # Row 10: Latitude
        cell_lat_r10 = ws.cell(row=label_lat_row, column=excel_data_col)
        cell_lat_r10.value = station_data_row.get("Latitude", "")
        cell_lat_r10.alignment = center_align_v_center_no_wrap

        # Row 11: Longitude
        cell_long_r11 = ws.cell(row=label_long_row, column=excel_data_col)
        cell_long_r11.value = station_data_row.get("Longitude", "")
        cell_long_r11.alignment = center_align_v_center_no_wrap
        # Row 12 (Optimum Static Values)
        # User's code uses "Static". Ensure this column exists in results_df.
        # If your allocate_slots provides "Optimum Static Param", change "Static" to that.
        opt_static_val = station_data_row.get("Static") # From user's code
        if opt_static_val is None: # Fallback if "Static" is not present
            opt_static_val = station_data_row.get("Optimum Static Param", "") # Check for my suggested name

        cell_opt_static = ws.cell(row=data_optimum_static_row, column=excel_data_col)
        cell_opt_static.value = opt_static_val
        cell_opt_static.alignment = center_align_v_center
        
        # Row 13 (Proposed Frequency Pair - Number and Color)
        frequency_val_r13 = station_data_row.get("Frequency")
        cell_freq_r13 = ws.cell(row=label_freq_pair_row, column=excel_data_col)
        if pd.notna(frequency_val_r13) and frequency_val_r13 != "N/A":
            try:
                frequency_r13 = int(frequency_val_r13)
                cell_freq_r13.value = frequency_r13
                bg_color_code_r13 = color_map.get(frequency_r13, "FFFFFF") # Default white
                cell_freq_r13.fill = PatternFill(start_color=bg_color_code_r13, end_color=bg_color_code_r13, fill_type="solid")
            except ValueError: 
                cell_freq_r13.value = str(frequency_val_r13) 
        else:
            cell_freq_r13.value = "N/A"
        cell_freq_r13.alignment = center_align_v_center

        # Row 14 (Stationary Count Data)
        cell_stat_count_r14 = ws.cell(row=stationary_count_row, column=excel_data_col)
        cell_stat_count_r14.value = station_data_row.get("Num Stationary Allocated", 0)
        cell_stat_count_r14.alignment = center_align_v_center
        
        # Row 15 (Stationary Kavach Tx Window - Slot Range)
        s_slots_str_r15 = str(station_data_row.get("Stationary Kavach Slots Allocated", ""))
        stat_slots_p_nums_r15 = [s.strip() for s in s_slots_str_r15.split(',') if s.strip() and s.strip().lower() != 'nan' and s.startswith('P')]
        tx_window_text_r15 = ""
        if stat_slots_p_nums_r15:
            slot_numbers_r15 = sorted([int(p[1:]) for p in stat_slots_p_nums_r15])
            if slot_numbers_r15: # Ensure list is not empty after parsing
                if len(slot_numbers_r15) == 1:
                    tx_window_text_r15 = f"P{slot_numbers_r15[0]}-P{slot_numbers_r15[0]+cell_stat_count_r14.value - 1}"
                else:
                    tx_window_text_r15 = f"P{slot_numbers_r15[0]}-P{slot_numbers_r15[-1]+cell_stat_count_r14.value - 1}"
        cell_tx_window_r15 = ws.cell(row=label_tx_window_row, column=excel_data_col)
        cell_tx_window_r15.value = tx_window_text_r15
        cell_tx_window_r15.alignment = center_align_v_center_no_wrap 

        # Row 16 (Onboard Count Data)
        cell_onboard_count_r16 = ws.cell(row=onboard_count_row, column=excel_data_col)
        cell_onboard_count_r16.value = station_data_row.get("Num Onboard Allocated", 0)
        cell_onboard_count_r16.alignment = center_align_v_center

    # --- Write Slot Names (P2, P3, ...) in Column 1 for the Matrix Data Area ---
    for r_idx_slot_name, slot_name in enumerate(all_slots):
        cell_slot_label = ws.cell(row=data_start_row + r_idx_slot_name, column=1)
        cell_slot_label.value = slot_name
        cell_slot_label.alignment = center_align_v_center

    # This is crucial for the adjacency check for the next slot
    # It should be the maximum 0-based index of your slots.
    # If all_slots = ["P2", "P3", ..., "P45"], len(all_slots) is 44. Max index is 43.
    if not all_slots: # Handle empty all_slots if it can occur
        # Depending on requirements, either return or raise error, or ensure all_slots is never empty
        print("Warning: all_slots is empty.")
        max_slot_idx_for_adj_check = -1 # Or handle appropriately
    else:
        max_slot_idx_for_adj_check = len(all_slots) - 1


    # --- Apply Styles, Colors, and Text to Matrix Data Cells (P-slot area) ---
    for r_idx_data_matrix, slot_in_current_row in enumerate(all_slots):
        current_excel_data_row = data_start_row + r_idx_data_matrix
        current_slot_0index = int(slot_in_current_row[1:]) - 2 # User's variable name

        for c_idx_df, station_name_for_coloring in enumerate(all_stations):
            excel_col_for_station = c_idx_df + 2
            cell_to_format = ws.cell(row=current_excel_data_row, column=excel_col_for_station)
            cell_to_format.value = "" 
            cell_to_format.font = Font() 
            cell_to_format.alignment = center_align_v_center

            station_data_rows_matrix = results_df[results_df["Station"] == station_name_for_coloring]
            if station_data_rows_matrix.empty: continue
            station_data_matrix_row = station_data_rows_matrix.iloc[0]

            s_slots_str_m = str(station_data_matrix_row.get("Stationary Kavach Slots Allocated", ""))
            o_p1_slots_str_m = str(station_data_matrix_row.get("Onboard Slots P1 Allocated", ""))
            o_p2_slots_str_m = str(station_data_matrix_row.get("Onboard Slots P2 Allocated", ""))
            o_p3_slots_str_m = str(station_data_matrix_row.get("Onboard Slots P3 Allocated", ""))
            o_p4_slots_str_m = str(station_data_matrix_row.get("Onboard Slots P4 Allocated", ""))
            o_p5_slots_str_m = str(station_data_matrix_row.get("Onboard Slots P5 Allocated", ""))
            o_p6_slots_str_m = str(station_data_matrix_row.get("Onboard Slots P6 Allocated", ""))


            set_stationary = {s.strip() for s in s_slots_str_m.split(',') if s.strip() and s.strip().lower() != 'nan'}
            set_onboard_p1 = {s.strip() for s in o_p1_slots_str_m.split(',') if s.strip() and s.strip().lower() != 'nan'}
            set_onboard_p2 = {s.strip() for s in o_p2_slots_str_m.split(',') if s.strip() and s.strip().lower() != 'nan'}
            set_onboard_p3 = {s.strip() for s in o_p3_slots_str_m.split(',') if s.strip() and s.strip().lower() != 'nan'}
            set_onboard_p4 = {s.strip() for s in o_p4_slots_str_m.split(',') if s.strip() and s.strip().lower() != 'nan'}
            set_onboard_p5 = {s.strip() for s in o_p5_slots_str_m.split(',') if s.strip() and s.strip().lower() != 'nan'}
            set_onboard_p6 = {s.strip() for s in o_p6_slots_str_m.split(',') if s.strip() and s.strip().lower() != 'nan'}



            is_stationary = slot_in_current_row in set_stationary
            is_onboard_p1 = slot_in_current_row in set_onboard_p1
            is_onboard_p2 = slot_in_current_row in set_onboard_p2
            is_onboard_p3 = slot_in_current_row in set_onboard_p3
            is_onboard_p4 = slot_in_current_row in set_onboard_p4
            is_onboard_p5 = slot_in_current_row in set_onboard_p5
            is_onboard_p6 = slot_in_current_row in set_onboard_p6

            
            if is_stationary:
                frequency_val_m = station_data_matrix_row.get("Frequency")
                if pd.notna(frequency_val_m) and frequency_val_m != "N/A":
                    try:
                        frequency_m = int(frequency_val_m)
                        bg_color_code_m = color_map.get(frequency_m, "FFFFFF")
                        cell_to_format.fill = PatternFill(start_color=bg_color_code_m, end_color=bg_color_code_m, fill_type="solid")
                    except ValueError: pass 
            
            if is_onboard_p1:
                cell_to_format.font = FONT_P1_STYLE
                cell_to_format.value = slot_in_current_row

            elif is_onboard_p2:
                cell_to_format.font = FONT_P2_STYLE
                cell_to_format.value = slot_in_current_row

            elif is_onboard_p3:
                cell_to_format.font = FONT_P3_STYLE
                cell_to_format.value = slot_in_current_row
   
            elif is_onboard_p4:
                cell_to_format.font = FONT_P4_STYLE
                cell_to_format.value = slot_in_current_row
  
            elif is_onboard_p5:
                cell_to_format.font = FONT_P5_STYLE
                cell_to_format.value = slot_in_current_row
  
            elif is_onboard_p6:
                cell_to_format.font = FONT_P6_STYLE
                cell_to_format.value = slot_in_current_row
   

            elif is_stationary: 
                cell_to_format.value = ""
            

    # --- Apply Borders to the Main Content Block ---
    # User's code: for r in range(7, max_excel_row):
    # max_excel_row in user's code was len(all_slots) + data_start_row -> 44 + 17 = 61
    # So loop is for r in range(7, 61) which means rows 7 to 60.
    # Row 60 is the last row of P-slot data.
    for r in range(label_stationary_kavach_id_row, (data_start_row + len(all_slots) -1) + 1):
        for c in range(1, max_excel_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = thin_border
            # Specific alignments have been set during content population.
            # No general alignment override here.

    # --- Add Allocation Details Sheet ---
    ws_details = wb.create_sheet(title="Allocation Details")
    if not results_df.empty:
        detail_headers = list(results_df.columns)
        ws_details.append(detail_headers)
        for _, row_data_obj_detail in results_df.iterrows():
            try:
                ws_details.append(list(row_data_obj_detail.astype(str))) 
            except Exception as e_row_detail:
                print(f"Warning: Could not append row to details sheet: {row_data_obj_detail.get('Station', 'Unknown_Station')}, Error: {e_row_detail}")
        
        for column_cells_detail in ws_details.columns:
            try:
                length = max(len(str(cell.value) if cell.value is not None else "") for cell in column_cells_detail)
                ws_details.column_dimensions[get_column_letter(column_cells_detail[0].column)].width = length + 2
            except Exception as e_col_size_detail:
                print(f"Warning: Could not resize column {get_column_letter(column_cells_detail[0].column)}, Error: {e_col_size_detail}")
    else:
        ws_details.append(["No allocation details to display."])

    try:
        wb.save(OUTPUT_FILE)
        print(f"Formatted and styled Excel file saved to: {OUTPUT_FILE}")
    except PermissionError:
        print(f"Critical Error: Permission denied. Failed to save Excel to {OUTPUT_FILE}.")
        print("Ensure the file is not open and you have write permissions.")
    except Exception as e_save:
        print(f"Critical Error: Failed to save Excel to {OUTPUT_FILE}. Error: {e_save}")

if __name__ == '__main__':
    stations = [
        {'name': 'LC.563', 'Static': 4, 'onboardSlots': 10, "StationCode": "LC563", "KavachID": "37023", "Latitude": 28.7041, "Longitude": 77.1025}, 
        {'name': 'Rundhi', 'Static': 4, 'onboardSlots': 14, "StationCode": "RUND", "KavachID": "37024", "Latitude": 29.0461, "Longitude": 76.90725}, 
        {'name': 'LC.560', 'Static': 4, 'onboardSlots': 9, "StationCode": "LC560", "KavachID": "37025", "Latitude": 31.7321, "Longitude": 74.4464},  
        {'name': 'Sholanka', 'Static': 4, 'onboardSlots': 16, "StationCode": "SHOL", "KavachID": "37026", "Latitude": 33.7041, "Longitude": 72.1025},
        {'name': 'LC.555', 'Static': 4, 'onboardSlots': 9, "StationCode": "LC555", "KavachID": "37027", "Latitude": 38.7041, "Longitude": 76.1025},  
        {'name': 'Hodal', 'Static': 4, 'onboardSlots': 14, "StationCode": "HODAL", "KavachID": "37028", "Latitude": 41.3041, "Longitude": 76.6647}, 
        {'name': 'Station G', 'Static': 6, 'onboardSlots': 20, "StationCode": "STG", "KavachID": "37029", "Latitude": 42.7055, "Longitude": 73.1025}, 
        {'name': 'Station H', 'Static': 15, 'onboardSlots': 35, "StationCode": "STH", "KavachID": "37030", "Latitude": 45.1041, "Longitude": 77.8272}, 
        {'name': 'Station I', 'Static': 5, 'onboardSlots': 18, "StationCode": "STI", "KavachID": "37031", "Latitude": 47.7041, "Longitude": 76.5278}, 
        {'name': 'Station J', 'Static': 8, 'onboardSlots': 5, "StationCode": "STJ", "KavachID": "37032", "Latitude": 49.4041, "Longitude": 76.8825}, 
        {'name': 'Station K', 'Static': 2, 'onboardSlots': 25, "StationCode": "STK", "KavachID": "37033", "Latitude": 51.7041, "Longitude": 77.6153}, 
        {'name': 'Station L', 'Static': 10, 'onboardSlots': 10, "StationCode": "STL", "KavachID": "37034", "Latitude": 54.3441, "Longitude": 73.1025},
        {'name': 'Mathura Jn', 'Static': 6, 'onboardSlots': 21, "StationCode": "MTHRJ", "KavachID": "37035", "Latitude": 57.5325, "Longitude": 73.6737},
    ]
    
    result_path = generate_excel(stations)

    if result_path:
        print(f"Process completed. Output file: {result_path}")
    else:
        print("Process failed.")